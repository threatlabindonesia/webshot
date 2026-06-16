#!/usr/bin/env python3
"""
webshot_checker.py

Domain/IP Screenshot Bulk Checker
- Single & bulk check dari file TXT
- Auto check HTTP + HTTPS per target
- Support format: domain, IP, domain:port, IP:port
- Screenshot disimpan dengan nama aman
- Output Excel (.xlsx) dan HTML report profesional
- Async + concurrent untuk performa tinggi (puluhan ribu target)
"""

import argparse
import asyncio
import base64
import hashlib
import os
import re
import sys
import time
from datetime import datetime
from pathlib import Path
from urllib.parse import urlparse
from zoneinfo import ZoneInfo

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from playwright.async_api import async_playwright, TimeoutError as PlaywrightTimeout

# ──────────────────────────────────────────────────────────────
# KONFIGURASI DEFAULT
# ──────────────────────────────────────────────────────────────
DEFAULT_TIMEOUT    = 15_000   # ms per request
DEFAULT_CONCURRENCY = 10      # browser contexts paralel
DEFAULT_SCREENSHOT_DIR = "screenshots"
DEFAULT_OUT_PREFIX = "webshot_report"
VIEWPORT = {"width": 1366, "height": 768}
TIMEZONE = "Asia/Jakarta"

STATUS_EMOJI = {
    "ok":      "✅",
    "error":   "❌",
    "timeout": "⏱️",
    "redirect":"↪️",
}

# ──────────────────────────────────────────────────────────────
# UTILITAS
# ──────────────────────────────────────────────────────────────

def normalize_target(raw: str) -> dict:
    """
    Parse target mentah: bisa domain, IP, domain:port, IP:port, atau URL penuh.
    Return dict berisi {original, host, port, schemes_to_try}
    """
    raw = raw.strip()
    if not raw or raw.startswith("#"):
        return None

    # Jika sudah ada scheme
    if raw.startswith("http://") or raw.startswith("https://"):
        parsed = urlparse(raw)
        return {
            "original": raw,
            "host": parsed.hostname,
            "port": parsed.port,
            "schemes_to_try": [raw],
        }

    # Pisahkan host dan port (handle IPv6 juga)
    # Format: host:port atau host
    port = None
    host = raw

    # IPv6 bracket notation [::1]:8080
    ipv6_port = re.match(r"^\[(.+)\]:(\d+)$", raw)
    if ipv6_port:
        host = f"[{ipv6_port.group(1)}]"
        port = ipv6_port.group(2)
    else:
        # domain:port atau IP:port (tapi bukan IPv6 tanpa bracket)
        parts = raw.rsplit(":", 1)
        if len(parts) == 2 and parts[1].isdigit():
            host = parts[0]
            port = parts[1]

    def build_url(scheme):
        if port:
            return f"{scheme}://{host}:{port}"
        return f"{scheme}://{host}"

    return {
        "original": raw,
        "host": host,
        "port": port,
        "schemes_to_try": [build_url("https"), build_url("http")],
    }


def safe_filename(url: str, max_len: int = 100) -> str:
    """Buat nama file aman dari URL."""
    name = re.sub(r"https?://", "", url)
    name = re.sub(r"[^\w\-.]", "_", name)
    name = re.sub(r"_+", "_", name).strip("_")
    if len(name) > max_len:
        h = hashlib.md5(url.encode()).hexdigest()[:8]
        name = name[:max_len - 9] + "_" + h
    return name


def now_str():
    return datetime.now(ZoneInfo(TIMEZONE)).strftime("%Y-%m-%d %H:%M:%S")


def load_targets(source: str | None, single: str | None) -> list[dict]:
    """Load target dari file atau argumen single."""
    raws = []

    if single:
        raws.append(single)

    if source:
        path = Path(source)
        if not path.exists():
            sys.exit(f"[ERROR] File tidak ditemukan: {path}")
        with path.open("r", encoding="utf-8", errors="ignore") as f:
            for line in f:
                line = line.strip()
                if line and not line.startswith("#"):
                    raws.append(line)

    seen = set()
    targets = []
    for raw in raws:
        t = normalize_target(raw)
        if t and raw not in seen:
            seen.add(raw)
            targets.append(t)

    return targets


# ──────────────────────────────────────────────────────────────
# SCREENSHOT ENGINE (ASYNC)
# ──────────────────────────────────────────────────────────────

async def capture_url(context, url: str, screenshot_path: str, timeout: int) -> dict:
    """Coba akses satu URL dan ambil screenshot."""
    result = {
        "url": url,
        "status_code": "",
        "title": "",
        "screenshot_file": "",
        "error": "",
        "status": "error",
        "response_time_ms": "",
        "final_url": "",
        "content_type": "",
    }

    page = None
    t0 = time.monotonic()

    try:
        page = await context.new_page()
        await page.set_viewport_size(VIEWPORT)

        response = await page.goto(
            url,
            timeout=timeout,
            wait_until="domcontentloaded",
        )

        elapsed = int((time.monotonic() - t0) * 1000)
        result["response_time_ms"] = elapsed

        if response:
            result["status_code"] = response.status
            result["final_url"] = page.url
            ct = response.headers.get("content-type", "")
            result["content_type"] = ct.split(";")[0].strip()

        # Tunggu sebentar agar render lebih sempurna
        try:
            await asyncio.wait_for(page.wait_for_load_state("networkidle"), timeout=3)
        except Exception:
            pass

        result["title"] = (await page.title()) or ""

        # Screenshot
        Path(screenshot_path).parent.mkdir(parents=True, exist_ok=True)
        await page.screenshot(path=screenshot_path, full_page=False)
        result["screenshot_file"] = Path(screenshot_path).name

        sc = result["status_code"]
        if isinstance(sc, int) and 300 <= sc < 400:
            result["status"] = "redirect"
        else:
            result["status"] = "ok"

    except PlaywrightTimeout:
        elapsed = int((time.monotonic() - t0) * 1000)
        result["response_time_ms"] = elapsed
        result["error"] = "TIMEOUT"
        result["status"] = "timeout"
    except Exception as e:
        elapsed = int((time.monotonic() - t0) * 1000)
        result["response_time_ms"] = elapsed
        result["error"] = str(e)[:200]
        result["status"] = "error"
    finally:
        if page:
            try:
                await page.close()
            except Exception:
                pass

    return result


async def process_target(semaphore, browser, target: dict, screenshot_dir: str, timeout: int) -> list[dict]:
    """Proses satu target (bisa HTTP + HTTPS)."""
    rows = []
    context = None

    async with semaphore:
        try:
            context = await browser.new_context(
                ignore_https_errors=True,
                user_agent=(
                    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                    "AppleWebKit/537.36 (KHTML, like Gecko) "
                    "Chrome/124.0.0.0 Safari/537.36"
                ),
            )

            for url in target["schemes_to_try"]:
                fname = safe_filename(url) + ".png"
                shot_path = str(Path(screenshot_dir) / fname)

                result = await capture_url(context, url, shot_path, timeout)
                result["original_target"] = target["original"]
                result["checked_at"] = now_str()

                emoji = STATUS_EMOJI.get(result["status"], "❓")
                sc = result["status_code"] or "-"
                rt = f"{result['response_time_ms']}ms" if result["response_time_ms"] else "-"
                print(
                    f"  {emoji} [{sc}] {url} | {rt}"
                    + (f" | {result['error']}" if result["error"] else "")
                )

                rows.append(result)

        except Exception as e:
            for url in target["schemes_to_try"]:
                rows.append({
                    "url": url,
                    "original_target": target["original"],
                    "status_code": "",
                    "title": "",
                    "screenshot_file": "",
                    "error": str(e)[:200],
                    "status": "error",
                    "response_time_ms": "",
                    "final_url": "",
                    "content_type": "",
                    "checked_at": now_str(),
                })
        finally:
            if context:
                try:
                    await context.close()
                except Exception:
                    pass

    return rows


async def run_scan(targets: list[dict], screenshot_dir: str, concurrency: int, timeout: int) -> list[dict]:
    """Main async scan loop."""
    all_rows = []
    semaphore = asyncio.Semaphore(concurrency)
    total = len(targets)

    print(f"\n[INFO] Mulai scan {total} target | concurrency={concurrency} | timeout={timeout}ms\n")

    async with async_playwright() as pw:
        browser = await pw.chromium.launch(
            headless=True,
            args=[
                "--no-sandbox",
                "--disable-setuid-sandbox",
                "--disable-dev-shm-usage",
                "--disable-gpu",
                "--disable-extensions",
            ],
        )

        tasks = [
            process_target(semaphore, browser, t, screenshot_dir, timeout)
            for t in targets
        ]

        done = 0
        for coro in asyncio.as_completed(tasks):
            rows = await coro
            all_rows.extend(rows)
            done += 1
            if done % 50 == 0 or done == total:
                pct = done / total * 100
                ok = sum(1 for r in all_rows if r["status"] == "ok")
                print(f"\n  📊 Progress: {done}/{total} ({pct:.1f}%) | OK: {ok}\n")

        await browser.close()

    return all_rows


# ──────────────────────────────────────────────────────────────
# EXCEL EXPORT
# ──────────────────────────────────────────────────────────────

COLUMNS_EXCEL = [
    ("original_target",  "Target",           20),
    ("url",              "URL Dicek",         40),
    ("status",           "Status",            12),
    ("status_code",      "HTTP Code",         12),
    ("title",            "Page Title",        40),
    ("screenshot_file",  "File Screenshot",   35),
    ("final_url",        "Final URL",         45),
    ("content_type",     "Content-Type",      25),
    ("response_time_ms", "Response Time (ms)", 18),
    ("error",            "Error",             40),
    ("checked_at",       "Waktu Check",       22),
]

STATUS_COLORS = {
    "ok":       "C8E6C9",
    "redirect": "FFF9C4",
    "timeout":  "FFE0B2",
    "error":    "FFCDD2",
}


def export_excel(rows: list[dict], output_path: str):
    wb = Workbook()
    ws = wb.active
    ws.title = "Hasil Scan"

    # Header
    header_fill = PatternFill("solid", fgColor="1A237E")
    header_font = Font(color="FFFFFF", bold=True, name="Calibri", size=11)

    for col_idx, (key, label, width) in enumerate(COLUMNS_EXCEL, 1):
        cell = ws.cell(row=1, column=col_idx, value=label)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.row_dimensions[1].height = 30
    ws.freeze_panes = "A2"

    # Data rows
    for row_idx, row in enumerate(rows, 2):
        fill_color = STATUS_COLORS.get(row.get("status", "error"), "FAFAFA")
        row_fill = PatternFill("solid", fgColor=fill_color)

        for col_idx, (key, label, width) in enumerate(COLUMNS_EXCEL, 1):
            val = row.get(key, "")
            if val is None:
                val = ""
            cell = ws.cell(row=row_idx, column=col_idx, value=str(val) if val != "" else "")
            cell.fill = row_fill
            cell.alignment = Alignment(vertical="center", wrap_text=False)
            if col_idx == 1:
                cell.font = Font(bold=True, name="Calibri")

        ws.row_dimensions[row_idx].height = 18

    # Summary sheet
    ws2 = wb.create_sheet("Ringkasan")
    total = len(rows)
    ok_count      = sum(1 for r in rows if r.get("status") == "ok")
    error_count   = sum(1 for r in rows if r.get("status") == "error")
    timeout_count = sum(1 for r in rows if r.get("status") == "timeout")
    redirect_count = sum(1 for r in rows if r.get("status") == "redirect")

    summary_data = [
        ("Total URL Dicek",  total),
        ("✅ Berhasil (OK)", ok_count),
        ("↪️ Redirect",       redirect_count),
        ("⏱️ Timeout",        timeout_count),
        ("❌ Error",          error_count),
        ("Success Rate",      f"{ok_count/total*100:.1f}%" if total else "0%"),
    ]

    ws2.column_dimensions["A"].width = 25
    ws2.column_dimensions["B"].width = 15
    ws2.cell(row=1, column=1, value="Metrik").font = Font(bold=True, size=12)
    ws2.cell(row=1, column=2, value="Nilai").font = Font(bold=True, size=12)

    for i, (label, val) in enumerate(summary_data, 2):
        ws2.cell(row=i, column=1, value=label)
        ws2.cell(row=i, column=2, value=val)

    wb.save(output_path)
    print(f"[OK] Excel disimpan: {output_path}")


# ──────────────────────────────────────────────────────────────
# HTML REPORT EXPORT
# ──────────────────────────────────────────────────────────────

def export_html(rows: list[dict], output_path: str, screenshot_dir: str):
    total = len(rows)
    ok_count      = sum(1 for r in rows if r.get("status") == "ok")
    error_count   = sum(1 for r in rows if r.get("status") == "error")
    timeout_count = sum(1 for r in rows if r.get("status") == "timeout")
    redirect_count = sum(1 for r in rows if r.get("status") == "redirect")
    success_rate = f"{ok_count/total*100:.1f}" if total else "0"
    generated_at = now_str()

    # Encode screenshots to base64 for self-contained HTML
    def shot_img_tag(filename):
        if not filename:
            return '<div class="no-shot">No Screenshot</div>'
        path = Path(screenshot_dir) / filename
        if path.exists():
            try:
                data = base64.b64encode(path.read_bytes()).decode()
                return f'<img src="data:image/png;base64,{data}" alt="{filename}" loading="lazy" onclick="openModal(this.src,\'{filename}\')">'
            except Exception:
                pass
        return f'<div class="no-shot">📁 {filename}</div>'

    status_badge = {
        "ok":       '<span class="badge badge-ok">✅ OK</span>',
        "redirect": '<span class="badge badge-redirect">↪️ Redirect</span>',
        "timeout":  '<span class="badge badge-timeout">⏱️ Timeout</span>',
        "error":    '<span class="badge badge-error">❌ Error</span>',
    }

    rows_html = ""
    for r in rows:
        st = r.get("status", "error")
        badge = status_badge.get(st, f'<span class="badge badge-error">{st}</span>')
        sc = r.get("status_code", "") or "-"
        rt = f"{r.get('response_time_ms', '')}ms" if r.get("response_time_ms") else "-"
        title = r.get("title", "") or "-"
        final_url = r.get("final_url", "") or ""
        error = r.get("error", "") or ""
        shot_fname = r.get("screenshot_file", "") or ""
        img_tag = shot_img_tag(shot_fname)
        orig = r.get("original_target", "")
        url = r.get("url", "")
        checked_at = r.get("checked_at", "")
        ct = r.get("content_type", "") or "-"

        rows_html += f"""
<tr class="row-{st}" data-status="{st}">
  <td class="col-target">
    <div class="target-orig">{orig}</div>
    <div class="target-url"><a href="{url}" target="_blank" rel="noopener">{url}</a></div>
  </td>
  <td class="col-status">{badge}<br><code>{sc}</code></td>
  <td class="col-title" title="{title}">{title[:80]}{"…" if len(title)>80 else ""}</td>
  <td class="col-shot">{img_tag}<div class="shot-name">{shot_fname or "—"}</div></td>
  <td class="col-meta">
    <div>⏱ {rt}</div>
    <div class="meta-ct">{ct}</div>
    {"<div class='meta-final'>→ " + final_url[:60] + "</div>" if final_url and final_url != url else ""}
    {"<div class='meta-error'>⚠ " + error[:80] + "</div>" if error else ""}
    <div class="meta-time">{checked_at}</div>
  </td>
</tr>"""

    html = f"""<!DOCTYPE html>
<html lang="id">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>WebShot Report — {generated_at}</title>
<style>
  @import url('https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700&family=JetBrains+Mono:wght@400;500&display=swap');

  :root {{
    --bg:        #0D1117;
    --surface:   #161B22;
    --surface2:  #21262D;
    --border:    #30363D;
    --text:      #E6EDF3;
    --muted:     #7D8590;
    --accent:    #58A6FF;
    --ok:        #3FB950;
    --warn:      #D29922;
    --error:     #F85149;
    --timeout:   #E3B341;
    --redirect:  #58A6FF;
    --radius:    8px;
  }}

  *, *::before, *::after {{ box-sizing: border-box; margin: 0; padding: 0; }}

  body {{
    font-family: 'Inter', system-ui, sans-serif;
    background: var(--bg);
    color: var(--text);
    min-height: 100vh;
    font-size: 13px;
    line-height: 1.5;
  }}

  /* ── HEADER ── */
  .header {{
    background: linear-gradient(135deg, #161B22 0%, #0D1117 100%);
    border-bottom: 1px solid var(--border);
    padding: 24px 32px;
    position: sticky;
    top: 0;
    z-index: 100;
    backdrop-filter: blur(10px);
  }}

  .header-inner {{
    max-width: 1600px;
    margin: 0 auto;
    display: flex;
    align-items: center;
    justify-content: space-between;
    gap: 24px;
    flex-wrap: wrap;
  }}

  .logo {{
    display: flex;
    align-items: center;
    gap: 12px;
  }}

  .logo-icon {{
    width: 36px; height: 36px;
    background: linear-gradient(135deg, #58A6FF, #BC8CFF);
    border-radius: 10px;
    display: flex; align-items: center; justify-content: center;
    font-size: 18px;
  }}

  .logo h1 {{
    font-size: 18px;
    font-weight: 700;
    letter-spacing: -0.3px;
    color: var(--text);
  }}

  .logo p {{
    font-size: 11px;
    color: var(--muted);
    font-family: 'JetBrains Mono', monospace;
  }}

  /* ── STAT CARDS ── */
  .stats {{
    display: flex;
    gap: 12px;
    flex-wrap: wrap;
  }}

  .stat {{
    background: var(--surface);
    border: 1px solid var(--border);
    border-radius: var(--radius);
    padding: 8px 16px;
    text-align: center;
    min-width: 80px;
  }}

  .stat-num {{
    font-size: 22px;
    font-weight: 700;
    font-family: 'JetBrains Mono', monospace;
    line-height: 1;
  }}

  .stat-label {{
    font-size: 10px;
    color: var(--muted);
    text-transform: uppercase;
    letter-spacing: 0.5px;
    margin-top: 2px;
  }}

  .stat-total  .stat-num {{ color: var(--accent); }}
  .stat-ok     .stat-num {{ color: var(--ok); }}
  .stat-error  .stat-num {{ color: var(--error); }}
  .stat-timeout .stat-num {{ color: var(--timeout); }}
  .stat-redirect .stat-num {{ color: var(--redirect); }}
  .stat-rate   .stat-num {{ color: #BC8CFF; }}

  /* ── TOOLBAR ── */
  .toolbar {{
    max-width: 1600px;
    margin: 16px auto 0;
    padding: 0 32px;
    display: flex;
    gap: 8px;
    flex-wrap: wrap;
    align-items: center;
  }}

  .filter-btn {{
    background: var(--surface);
    border: 1px solid var(--border);
    color: var(--muted);
    padding: 6px 14px;
    border-radius: 20px;
    cursor: pointer;
    font-size: 12px;
    font-family: inherit;
    transition: all .15s;
  }}

  .filter-btn:hover, .filter-btn.active {{
    background: var(--accent);
    border-color: var(--accent);
    color: #000;
    font-weight: 600;
  }}

  .search-box {{
    margin-left: auto;
    background: var(--surface);
    border: 1px solid var(--border);
    color: var(--text);
    padding: 6px 14px;
    border-radius: 20px;
    font-size: 12px;
    font-family: inherit;
    width: 240px;
    outline: none;
    transition: border-color .15s;
  }}

  .search-box:focus {{ border-color: var(--accent); }}
  .search-box::placeholder {{ color: var(--muted); }}

  /* ── TABLE ── */
  .table-wrap {{
    max-width: 1600px;
    margin: 16px auto 40px;
    padding: 0 32px;
    overflow-x: auto;
  }}

  table {{
    width: 100%;
    border-collapse: separate;
    border-spacing: 0 4px;
  }}

  thead th {{
    background: var(--surface);
    color: var(--muted);
    font-size: 11px;
    text-transform: uppercase;
    letter-spacing: 0.6px;
    font-weight: 600;
    padding: 10px 12px;
    border-top: 1px solid var(--border);
    border-bottom: 1px solid var(--border);
    white-space: nowrap;
    position: sticky;
    top: 100px;
    z-index: 10;
  }}

  thead th:first-child {{ border-left: 1px solid var(--border); border-radius: 6px 0 0 6px; }}
  thead th:last-child  {{ border-right: 1px solid var(--border); border-radius: 0 6px 6px 0; }}

  tbody tr {{
    background: var(--surface);
    transition: background .12s;
  }}

  tbody tr:hover {{ background: var(--surface2); }}

  td {{
    padding: 10px 12px;
    vertical-align: middle;
    border-top: 1px solid var(--border);
    border-bottom: 1px solid var(--border);
  }}

  td:first-child {{ border-left: 1px solid var(--border); border-radius: 6px 0 0 6px; }}
  td:last-child  {{ border-right: 1px solid var(--border); border-radius: 0 6px 6px 0; }}

  /* ── STATUS BADGES ── */
  .badge {{
    display: inline-block;
    padding: 2px 8px;
    border-radius: 12px;
    font-size: 11px;
    font-weight: 600;
    white-space: nowrap;
  }}

  .badge-ok       {{ background: rgba(63,185,80,.15);  color: var(--ok); }}
  .badge-error    {{ background: rgba(248,81,73,.15);  color: var(--error); }}
  .badge-timeout  {{ background: rgba(227,179,65,.15); color: var(--timeout); }}
  .badge-redirect {{ background: rgba(88,166,255,.15); color: var(--redirect); }}

  /* ── COLUMNS ── */
  .col-target {{ min-width: 250px; max-width: 320px; }}
  .col-status {{ min-width: 90px; text-align: center; }}
  .col-title  {{ min-width: 200px; max-width: 280px; color: var(--muted); }}
  .col-shot   {{ min-width: 160px; max-width: 200px; text-align: center; }}
  .col-meta   {{ min-width: 180px; font-size: 11px; color: var(--muted); }}

  .target-orig {{ font-family: 'JetBrains Mono', monospace; font-size: 12px; font-weight: 500; color: var(--text); }}
  .target-url  {{ font-size: 11px; color: var(--muted); margin-top: 2px; }}
  .target-url a {{ color: var(--accent); text-decoration: none; }}
  .target-url a:hover {{ text-decoration: underline; }}

  .col-shot img {{
    max-width: 180px;
    max-height: 100px;
    border-radius: 4px;
    border: 1px solid var(--border);
    cursor: zoom-in;
    display: block;
    margin: 0 auto 4px;
    transition: transform .2s, box-shadow .2s;
  }}

  .col-shot img:hover {{
    transform: scale(1.05);
    box-shadow: 0 4px 20px rgba(0,0,0,.5);
  }}

  .no-shot {{
    background: var(--surface2);
    border: 1px dashed var(--border);
    border-radius: 4px;
    padding: 16px;
    color: var(--muted);
    font-size: 11px;
    text-align: center;
  }}

  .shot-name {{
    font-family: 'JetBrains Mono', monospace;
    font-size: 9px;
    color: var(--muted);
    margin-top: 3px;
    word-break: break-all;
  }}

  .meta-error {{ color: var(--error); margin-top: 3px; }}
  .meta-final {{ color: var(--accent); font-size: 10px; margin-top: 2px; }}
  .meta-ct    {{ color: var(--muted); font-size: 10px; margin-top: 1px; }}
  .meta-time  {{ color: var(--muted); font-size: 10px; margin-top: 4px; }}

  code {{
    background: var(--surface2);
    padding: 1px 5px;
    border-radius: 4px;
    font-family: 'JetBrains Mono', monospace;
    font-size: 11px;
    color: var(--muted);
    margin-top: 4px;
    display: inline-block;
  }}

  /* ── MODAL ── */
  #modal {{
    display: none;
    position: fixed; inset: 0;
    background: rgba(0,0,0,.9);
    z-index: 1000;
    align-items: center;
    justify-content: center;
    flex-direction: column;
    gap: 12px;
  }}

  #modal.open {{ display: flex; }}

  #modal img {{
    max-width: 90vw;
    max-height: 80vh;
    border-radius: var(--radius);
    border: 1px solid var(--border);
    box-shadow: 0 20px 60px rgba(0,0,0,.8);
  }}

  #modal-label {{
    color: var(--muted);
    font-family: 'JetBrains Mono', monospace;
    font-size: 12px;
  }}

  #modal-close {{
    position: absolute; top: 20px; right: 24px;
    background: var(--surface);
    border: 1px solid var(--border);
    color: var(--text);
    font-size: 20px;
    width: 36px; height: 36px;
    border-radius: 50%;
    cursor: pointer;
    display: flex; align-items: center; justify-content: center;
    line-height: 1;
  }}

  /* ── HIDDEN ── */
  tr.hidden {{ display: none; }}

  /* ── FOOTER ── */
  .footer {{
    text-align: center;
    color: var(--muted);
    font-size: 11px;
    padding: 24px;
    border-top: 1px solid var(--border);
  }}
</style>
</head>
<body>

<header class="header">
  <div class="header-inner">
    <div class="logo">
      <div class="logo-icon">🔍</div>
      <div>
        <h1>WebShot Report</h1>
        <p>Generated: {generated_at} WIB</p>
      </div>
    </div>
    <div class="stats">
      <div class="stat stat-total">
        <div class="stat-num">{total}</div>
        <div class="stat-label">Total</div>
      </div>
      <div class="stat stat-ok">
        <div class="stat-num">{ok_count}</div>
        <div class="stat-label">OK</div>
      </div>
      <div class="stat stat-redirect">
        <div class="stat-num">{redirect_count}</div>
        <div class="stat-label">Redirect</div>
      </div>
      <div class="stat stat-timeout">
        <div class="stat-num">{timeout_count}</div>
        <div class="stat-label">Timeout</div>
      </div>
      <div class="stat stat-error">
        <div class="stat-num">{error_count}</div>
        <div class="stat-label">Error</div>
      </div>
      <div class="stat stat-rate">
        <div class="stat-num">{success_rate}%</div>
        <div class="stat-label">Success</div>
      </div>
    </div>
  </div>
</header>

<div class="toolbar">
  <button class="filter-btn active" onclick="filterStatus('all', this)">All ({total})</button>
  <button class="filter-btn" onclick="filterStatus('ok', this)">✅ OK ({ok_count})</button>
  <button class="filter-btn" onclick="filterStatus('redirect', this)">↪️ Redirect ({redirect_count})</button>
  <button class="filter-btn" onclick="filterStatus('timeout', this)">⏱️ Timeout ({timeout_count})</button>
  <button class="filter-btn" onclick="filterStatus('error', this)">❌ Error ({error_count})</button>
  <input class="search-box" type="search" placeholder="🔍 Cari domain / URL / title..." oninput="searchTable(this.value)">
</div>

<div class="table-wrap">
  <table id="main-table">
    <thead>
      <tr>
        <th>Target / URL</th>
        <th>Status</th>
        <th>Page Title</th>
        <th>Screenshot</th>
        <th>Info</th>
      </tr>
    </thead>
    <tbody id="tbody">
      {rows_html}
    </tbody>
  </table>
</div>

<div class="footer">
  WebShot Checker &mdash; {total} URLs scanned &mdash; {generated_at}
</div>

<!-- MODAL -->
<div id="modal" onclick="closeModal()">
  <button id="modal-close" onclick="closeModal()">✕</button>
  <img id="modal-img" src="" alt="">
  <div id="modal-label"></div>
</div>

<script>
  let currentFilter = 'all';
  let currentSearch = '';

  function applyFilters() {{
    const rows = document.querySelectorAll('#tbody tr');
    const q = currentSearch.toLowerCase();
    rows.forEach(tr => {{
      const matchStatus = currentFilter === 'all' || tr.dataset.status === currentFilter;
      const text = tr.textContent.toLowerCase();
      const matchSearch = !q || text.includes(q);
      tr.classList.toggle('hidden', !(matchStatus && matchSearch));
    }});
  }}

  function filterStatus(status, btn) {{
    currentFilter = status;
    document.querySelectorAll('.filter-btn').forEach(b => b.classList.remove('active'));
    btn.classList.add('active');
    applyFilters();
  }}

  function searchTable(val) {{
    currentSearch = val;
    applyFilters();
  }}

  function openModal(src, name) {{
    document.getElementById('modal-img').src = src;
    document.getElementById('modal-label').textContent = name;
    document.getElementById('modal').classList.add('open');
  }}

  function closeModal() {{
    document.getElementById('modal').classList.remove('open');
  }}

  document.addEventListener('keydown', e => {{
    if (e.key === 'Escape') closeModal();
  }});
</script>
</body>
</html>"""

    Path(output_path).write_text(html, encoding="utf-8")
    print(f"[OK] HTML Report disimpan: {output_path}")


# ──────────────────────────────────────────────────────────────
# MAIN
# ──────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(
        description="🔍 WebShot Checker — Bulk domain/IP screenshot scanner",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Contoh penggunaan:
  # Single target
  python webshot_checker.py --target google.com

  # Bulk dari file
  python webshot_checker.py --file targets.txt

  # Dengan port kustom
  python webshot_checker.py --target 192.168.1.1:8080

  # Bulk + filter output
  python webshot_checker.py --file targets.txt --out report --format both --concurrency 20

Format file targets.txt (satu per baris):
  google.com
  192.168.1.1
  example.com:8443
  10.0.0.1:9090
  https://custom.domain/path
        """
    )

    parser.add_argument("--target",      help="Single target (domain, IP, atau IP:port)")
    parser.add_argument("--file",        help="File TXT berisi daftar target (satu per baris)")
    parser.add_argument("--out",         default=DEFAULT_OUT_PREFIX, help="Prefix nama file output")
    parser.add_argument("--format",      choices=["excel", "html", "both"], default="both",
                        help="Format output: excel | html | both")
    parser.add_argument("--shots-dir",   default=DEFAULT_SCREENSHOT_DIR,
                        help=f"Folder penyimpanan screenshot (default: {DEFAULT_SCREENSHOT_DIR})")
    parser.add_argument("--concurrency", type=int, default=DEFAULT_CONCURRENCY,
                        help=f"Jumlah tab paralel (default: {DEFAULT_CONCURRENCY})")
    parser.add_argument("--timeout",     type=int, default=DEFAULT_TIMEOUT // 1000,
                        help=f"Timeout per request dalam detik (default: {DEFAULT_TIMEOUT//1000})")
    parser.add_argument("--no-http",     action="store_true",
                        help="Skip check HTTP, hanya HTTPS")
    parser.add_argument("--no-https",    action="store_true",
                        help="Skip check HTTPS, hanya HTTP")

    args = parser.parse_args()

    if not args.target and not args.file:
        parser.error("Harap tentukan --target atau --file")

    targets = load_targets(args.file, args.target)

    if not targets:
        sys.exit("[ERROR] Tidak ada target valid ditemukan.")

    # Filter scheme jika diminta
    if args.no_http or args.no_https:
        for t in targets:
            if args.no_http:
                t["schemes_to_try"] = [u for u in t["schemes_to_try"] if not u.startswith("http://")]
            if args.no_https:
                t["schemes_to_try"] = [u for u in t["schemes_to_try"] if not u.startswith("https://")]
        targets = [t for t in targets if t["schemes_to_try"]]

    # Buat folder screenshot
    Path(args.shots_dir).mkdir(parents=True, exist_ok=True)

    print(f"""
╔══════════════════════════════════════════════════════╗
║         🔍  WebShot Checker — Bulk Scanner           ║
╠══════════════════════════════════════════════════════╣
║  Target       : {len(targets):<35} ║
║  Concurrency  : {args.concurrency:<35} ║
║  Timeout      : {args.timeout}s{'':<33} ║
║  Screenshots  : {args.shots_dir:<35} ║
║  Output       : {args.out:<35} ║
╚══════════════════════════════════════════════════════╝
""")

    t_start = time.monotonic()

    rows = asyncio.run(
        run_scan(
            targets=targets,
            screenshot_dir=args.shots_dir,
            concurrency=args.concurrency,
            timeout=args.timeout * 1000,
        )
    )

    elapsed = time.monotonic() - t_start
    ok = sum(1 for r in rows if r["status"] == "ok")

    print(f"\n[DONE] Selesai dalam {elapsed:.1f}s | {len(rows)} URL | {ok} berhasil\n")

    # Export
    if args.format in ("excel", "both"):
        export_excel(rows, f"{args.out}.xlsx")

    if args.format in ("html", "both"):
        export_html(rows, f"{args.out}.html", args.shots_dir)

    print("\n✅ Semua selesai!")


if __name__ == "__main__":
    main()
